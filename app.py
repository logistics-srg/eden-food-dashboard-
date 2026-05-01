import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import plotly.express as px
import base64
import os
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from io import BytesIO

st.set_page_config(page_title="EDEN FOOD", page_icon="🍌",
                   layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');
header[data-testid="stHeader"],#MainMenu,.stAppDeployButton,footer{display:none!important}
*{font-family:'Inter',sans-serif!important;box-sizing:border-box}
.block-container{padding:0!important;max-width:100%!important}
.stApp{background:#F4F6FB!important}
:root{
  --primary:#4361EE;--primary-light:#EEF2FF;--success:#10B981;--success-light:#D1FAE5;
  --warning:#F59E0B;--warning-light:#FEF3C7;--danger:#EF4444;--danger-light:#FEE2E2;
  --info:#3B82F6;--info-light:#DBEAFE;--surface:#FFFFFF;--surface-2:#F9FAFB;
  --border:#E5E7EB;--text-primary:#111827;--text-secondary:#6B7280;--text-muted:#9CA3AF;
  --shadow-sm:0 1px 2px rgba(0,0,0,0.05);
  --shadow-md:0 4px 6px -1px rgba(0,0,0,0.07),0 2px 4px -1px rgba(0,0,0,0.04);
  --shadow-lg:0 10px 15px -3px rgba(0,0,0,0.08),0 4px 6px -2px rgba(0,0,0,0.04);
  --radius-sm:8px;--radius-md:12px;--radius-lg:16px;--radius-xl:24px;--radius-full:9999px;
  --transition:all 0.2s cubic-bezier(0.4,0,0.2,1);
}
section[data-testid="stSidebar"]{
  background:#111111!important;border-right:1px solid rgba(255,255,255,0.06)!important;
  box-shadow:none!important;min-width:220px!important;max-width:220px!important;
}
section[data-testid="stSidebarCollapsedControl"]{display:none!important;width:0!important}
[data-testid="collapsedControl"]{display:none!important}
section[data-testid="stSidebar"] button[data-testid="baseButton-headerNoPadding"]{display:none!important}
section[data-testid="stSidebar"] *{font-family:'Inter',sans-serif!important}
section[data-testid="stSidebar"] .stButton>button{
  background:transparent!important;border:none!important;color:rgba(255,255,255,0.5)!important;
  text-align:left!important;width:100%!important;padding:9px 14px!important;
  border-radius:8px!important;font-size:13px!important;font-weight:500!important;
  transition:all 0.18s ease!important;margin-bottom:2px!important;
}
section[data-testid="stSidebar"] .stButton>button:hover{
  background:rgba(255,255,255,0.07)!important;color:#fff!important;
}
.main-wrap{padding:28px 36px;max-width:1500px;margin:0 auto}
.kpi-grid{display:grid;grid-template-columns:repeat(6,1fr);gap:14px;margin-bottom:28px}
.kpi-card{background:var(--surface);border-radius:var(--radius-lg);padding:20px 18px;
  border:1px solid var(--border);box-shadow:var(--shadow-sm);position:relative;overflow:hidden;
  transition:var(--transition);cursor:default;}
.kpi-card::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;
  border-radius:var(--radius-lg) var(--radius-lg) 0 0;
  background:linear-gradient(90deg,var(--card-color,#4361EE),transparent);}
.kpi-card:hover{transform:translateY(-3px);box-shadow:var(--shadow-lg)}
.kpi-card .icon-wrap{width:38px;height:38px;border-radius:10px;display:flex;align-items:center;
  justify-content:center;font-size:18px;margin-bottom:12px;background:var(--card-bg,var(--primary-light));}
.kpi-card .kpi-val{font-size:2rem;font-weight:900;letter-spacing:-1.5px;line-height:1;margin-bottom:3px}
.kpi-card .kpi-lbl{font-size:10px;color:var(--text-muted);text-transform:uppercase;letter-spacing:1px;font-weight:700;margin-bottom:4px}
.kpi-card .kpi-sub{font-size:11px;color:var(--text-muted)}
.sec-hdr{display:flex;align-items:center;justify-content:space-between;margin:28px 0 14px;
  padding-bottom:12px;border-bottom:1px solid var(--border);}
.sec-title{font-size:14px;font-weight:700;color:var(--text-primary)}
.sec-sub{font-size:12px;color:var(--text-muted)}
.cmd-row{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-md);
  padding:14px 18px;margin-bottom:6px;display:flex;align-items:center;
  justify-content:space-between;flex-wrap:wrap;gap:10px;transition:var(--transition);position:relative;}
.cmd-row::before{content:'';position:absolute;left:0;top:50%;transform:translateY(-50%);
  width:3px;height:0;border-radius:0 2px 2px 0;background:var(--primary);
  transition:all 0.35s cubic-bezier(0.4,0,0.2,1);}
.cmd-row:hover{box-shadow:var(--shadow-md);border-color:rgba(67,97,238,0.2);transform:translateX(2px)}
.cmd-row:hover::before{height:60%}
.pill{display:inline-flex;align-items:center;gap:5px;padding:4px 12px;
  border-radius:var(--radius-full);font-size:11px;font-weight:700;letter-spacing:0.2px}
.pill-green{background:var(--success-light);color:#065F46}
.pill-orange{background:var(--warning-light);color:#92400E}
.pill-red{background:var(--danger-light);color:#991B1B}
.pill-blue{background:var(--info-light);color:#1E40AF}
.card{background:var(--surface);border-radius:var(--radius-lg);padding:22px 24px;
  border:1px solid var(--border);box-shadow:var(--shadow-sm);margin-bottom:12px;transition:var(--transition);}
.card:hover{box-shadow:var(--shadow-md)}
.hero-wrap{position:relative;width:100%;height:320px;overflow:hidden;}
.hero-overlay{position:absolute;inset:0;
  background:linear-gradient(110deg,rgba(10,15,30,0.88) 0%,rgba(67,97,238,0.42) 55%,rgba(0,0,0,0.1) 100%);
  display:flex;align-items:center;padding:0 52px;}
.hero-text h1{font-size:34px;font-weight:900;color:#fff;letter-spacing:-1.2px;line-height:1.15;margin:0 0 10px}
.hero-text p{font-size:14px;color:rgba(255,255,255,0.78);max-width:440px;line-height:1.65;margin:0 0 22px}
.hero-badge{display:inline-flex;align-items:center;gap:7px;background:rgba(255,255,255,0.12);
  backdrop-filter:blur(12px);border:1px solid rgba(255,255,255,0.22);
  color:#fff;padding:7px 18px;border-radius:var(--radius-full);font-size:12px;font-weight:600;}
.alert{padding:11px 15px;border-radius:var(--radius-sm);font-size:12px;font-weight:600;margin:6px 0}
.alert-ok{background:var(--success-light);border-left:3px solid var(--success);color:#065F46}
.alert-warn{background:var(--warning-light);border-left:3px solid var(--warning);color:#92400E}
.alert-red{background:var(--danger-light);border-left:3px solid var(--danger);color:#991B1B}
.prog-track{background:#E5E7EB;border-radius:var(--radius-full);height:5px;overflow:hidden}
.prog-fill{height:100%;border-radius:var(--radius-full)}
.doc-zone{background:var(--surface-2);border:1.5px dashed #D1D5DB;border-radius:var(--radius-md);padding:18px}
.doc-chip{display:inline-flex;align-items:center;gap:5px;background:#fff;border:1px solid var(--border);
  border-radius:6px;padding:5px 10px;font-size:11px;color:var(--text-secondary);font-weight:500;margin:3px;}
div[data-testid="stForm"]{background:var(--surface)!important;border-radius:var(--radius-lg)!important;
  padding:26px!important;border:1px solid var(--border)!important;box-shadow:none!important;}
div[data-testid="stDataFrame"]{border-radius:var(--radius-md)!important;overflow:hidden!important}
hr{border:none!important;border-top:1px solid var(--border)!important;margin:16px 0!important}
::-webkit-scrollbar{width:5px;height:5px}
::-webkit-scrollbar-track{background:transparent}
::-webkit-scrollbar-thumb{background:#D1D5DB;border-radius:var(--radius-full)}
.track-card{background:var(--surface);border-radius:var(--radius-lg);border:1px solid var(--border);
  box-shadow:var(--shadow-sm);margin-bottom:14px;overflow:hidden;transition:var(--transition);}
.track-card:hover{box-shadow:var(--shadow-lg);transform:translateY(-2px)}
.track-header{padding:18px 22px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:10px}
.stage-dot{width:30px;height:30px;border-radius:50%;display:flex;align-items:center;justify-content:center;
  font-size:12px;transition:all 0.4s cubic-bezier(0.34,1.56,0.64,1);flex-shrink:0}
.stage-line{flex:1;height:2px;min-width:24px;position:relative;overflow:hidden}
.stage-line-fill{position:absolute;left:0;top:0;height:100%;border-radius:2px;transition:width 1.2s cubic-bezier(0.4,0,0.2,1)}
@keyframes shipPulse{0%,100%{transform:scale(1)}50%{transform:scale(1.15)}}
.ship-active{animation:shipPulse 2s ease-in-out infinite}
@keyframes dotPulse{0%,100%{opacity:1;box-shadow:0 0 6px #10B981}50%{opacity:0.5;box-shadow:0 0 3px #10B981}}
.status-dot{animation:dotPulse 2s ease-in-out infinite}
</style>
""", unsafe_allow_html=True)

# ── USERS ──────────────────────────────────────────────────────────────────────
USERS = {
    "yann":     {"password": "EdenFood2026!",  "role": "admin"},
    "eden":     {"password": "Eden@Logistik",   "role": "user"},
    "srg":      {"password": "SRG@Trading1",    "role": "user"},
    "edenfood": {"password": "Edenfood@2026",   "role": "user"},
}

for k, v in {
    "authenticated": False, "username": "", "role": "",
    "page": "dashboard", "new_commandes": [],
    "expanded_cmd": None, "expanded_track": set()
}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── HELPERS ────────────────────────────────────────────────────────────────────
def img_to_b64(path):
    try:
        ext  = path.split(".")[-1].lower()
        mime = "jpeg" if ext in ["jpg", "jpeg"] else "png"
        return "data:image/" + mime + ";base64," + base64.b64encode(open(path, "rb").read()).decode()
    except:
        return None

def apply_chart_style(fig, bgcolor="#fff"):
    fig.update_layout(
        paper_bgcolor=bgcolor, plot_bgcolor=bgcolor,
        font=dict(family="Inter", size=12, color="#111827"),
        margin=dict(t=20, b=20, l=20, r=20),
        legend=dict(font=dict(color="#111827", size=12))
    )
    fig.update_xaxes(tickfont=dict(color="#6B7280"), showgrid=False, linecolor="#E5E7EB")
    fig.update_yaxes(tickfont=dict(color="#6B7280"), gridcolor="#F3F4F6", linecolor="#F3F4F6")
    return fig

DOCS_ROOT = "docs"
DOC_TYPES = ["BL Draft", "BL Original", "Phytosanitaire", "Certificat Origine",
             "Facture Proforma", "Packing List", "Licence DPVCT", "Autre"]

def docs_path(booking):
    p = os.path.join(DOCS_ROOT, str(booking).replace("/", "_").replace(" ", "_"))
    os.makedirs(p, exist_ok=True)
    return p

def list_docs(booking):
    p = docs_path(booking)
    return [f for f in os.listdir(p) if not f.startswith(".")] if os.path.exists(p) else []

def save_doc(booking, uploaded_file):
    fp = os.path.join(docs_path(booking), uploaded_file.name)
    with open(fp, "wb") as f:
        f.write(uploaded_file.getbuffer())

def delete_doc(booking, filename):
    fp = os.path.join(docs_path(booking), filename)
    if os.path.exists(fp):
        os.remove(fp)

# ── ROUTING ────────────────────────────────────────────────────────────────────
ROUTING = {
    "MOIN(COSTA RICA)": {
        "total_days": 21,
        "stages": [
            {"id": "moin",       "icon": "⚓", "label": "MOIN CR",       "days": 0,  "color": "#F59E0B"},
            {"id": "caraibes",   "icon": "🌊", "label": "Caraibes",      "days": 3,  "color": "#3B82F6"},
            {"id": "atlantique", "icon": "🌊", "label": "Atlantique",    "days": 8,  "color": "#2563EB"},
            {"id": "algeciras",  "icon": "🔄", "label": "Algeciras",     "days": 15, "color": "#8B5CF6"},
            {"id": "alboran",    "icon": "🌊", "label": "Mer Alboran",   "days": 18, "color": "#0EA5E9"},
            {"id": "ghazaouet",  "icon": "⚓", "label": "Ghazaouet DZ", "days": 21, "color": "#10B981"},
        ]
    },
    "TURBO(COLOMBIA)": {
        "total_days": 19,
        "stages": [
            {"id": "turbo",      "icon": "⚓", "label": "TURBO COL",    "days": 0,  "color": "#F59E0B"},
            {"id": "caraibes",   "icon": "🌊", "label": "Caraibes",     "days": 2,  "color": "#3B82F6"},
            {"id": "atlantique", "icon": "🌊", "label": "Atlantique",   "days": 6,  "color": "#2563EB"},
            {"id": "algeciras",  "icon": "🔄", "label": "Algeciras",    "days": 13, "color": "#8B5CF6"},
            {"id": "alboran",    "icon": "🌊", "label": "Mer Alboran",  "days": 16, "color": "#0EA5E9"},
            {"id": "ghazaouet",  "icon": "⚓", "label": "Ghazaouet DZ","days": 19, "color": "#10B981"},
        ]
    }
}

def get_routing(pol):
    pol_up = str(pol).upper()
    for key in ROUTING:
        if key in pol_up or pol_up in key:
            return ROUTING[key]
    return ROUTING["MOIN(COSTA RICA)"]

def parse_depart(depart_str):
    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"]:
        try:
            return datetime.strptime(str(depart_str).strip(), fmt).date()
        except:
            continue
    return None

def get_tracking_info(depart_str, pol="MOIN(COSTA RICA)"):
    dep = parse_depart(depart_str)
    if not dep:
        return 0, 0.0, None, None
    routing    = get_routing(pol)
    total_days = routing["total_days"]
    stages     = routing["stages"]
    today      = date.today()
    days_el    = (today - dep).days
    eta        = dep + timedelta(days=total_days)
    stage_idx  = 0
    for i, s in enumerate(stages):
        if days_el >= s["days"]:
            stage_idx = i
    if days_el < 0:
        progress = 0.0
    elif days_el >= total_days:
        progress = 100.0
    else:
        progress = round(min(100.0, max(0.0, days_el / total_days * 100)), 1)
    return stage_idx, progress, days_el, eta

def build_stage_timeline_html(stage_idx, progress, pol):
    routing = get_routing(pol)
    stages  = routing["stages"]
    html = '<div style="display:flex;align-items:center;padding:14px 0;gap:0;min-width:580px;overflow-x:auto">'
    for i, s in enumerate(stages):
        is_done    = i < stage_idx
        is_current = i == stage_idx
        is_future  = i > stage_idx
        if is_done:
            dot_bg = "#10B981"; dot_border = "#10B981"; dot_shadow = ""
            dot_content = "v"; lbl_color = "#10B981"; lbl_weight = "600"
        elif is_current:
            dot_bg = s["color"]; dot_border = s["color"]
            dot_shadow = "box-shadow:0 0 0 4px " + s["color"] + "33,0 0 12px " + s["color"] + "55"
            dot_content = s["icon"]; lbl_color = s["color"]; lbl_weight = "800"
        else:
            dot_bg = "#F9FAFB"; dot_border = "#E5E7EB"; dot_shadow = ""
            dot_content = s["icon"]; lbl_color = "#9CA3AF"; lbl_weight = "500"
        anim      = 'class="ship-active"' if is_current and 0 < i < len(stages) - 1 else ""
        txt_color = "#fff" if not is_future else "#9CA3AF"
        fs        = "11px" if is_done else "14px"
        html += (
            '<div style="display:flex;flex-direction:column;align-items:center;gap:6px;flex-shrink:0">'
            + '<div ' + anim + ' class="stage-dot" style="background:' + dot_bg
            + ';border:2px solid ' + dot_border + ';' + dot_shadow
            + ';color:' + txt_color + ';font-size:' + fs + '">' + dot_content + '</div>'
            + '<div style="font-size:9px;color:' + lbl_color + ';font-weight:' + lbl_weight
            + ';text-align:center;max-width:72px;line-height:1.3">' + s["label"] + '</div>'
            + '</div>'
        )
        if i < len(stages) - 1:
            lc = "#10B981" if is_done else ("#E5E7EB" if is_future else s["color"])
            lf = 100 if is_done else 0
            html += (
                '<div class="stage-line" style="background:#E5E7EB;margin-bottom:16px">'
                + '<div class="stage-line-fill" style="background:' + lc + ';width:' + str(lf) + '%"></div></div>'
            )
    html += "</div>"
    return html

def render_route_map_svg(active_ships):
    WP = [(60, 150), (165, 140), (315, 115), (480, 92), (590, 78), (660, 88), (730, 98)]
    path_d = "M " + str(WP[0][0]) + " " + str(WP[0][1])
    for i in range(1, len(WP)):
        x0, y0 = WP[i - 1]; x1, y1 = WP[i]; cx = (x0 + x1) / 2
        path_d += " C " + str(cx) + " " + str(y0) + ", " + str(cx) + " " + str(y1) + ", " + str(x1) + " " + str(y1)
    fdefs = ""; ships = ""
    for idx, ship in enumerate(active_ships):
        fid = "sg" + str(idx); pct = max(0.0, min(1.0, ship["progress"] / 100))
        segs = len(WP) - 1; sf = pct * segs; si = min(int(sf), segs - 1); st_ = sf - si
        x0, y0 = WP[si]; x1, y1 = WP[min(si + 1, segs)]
        sx = round(x0 + (x1 - x0) * st_, 1); sy = round(y0 + (y1 - y0) * st_, 1)
        lbl = ship["label"][:9]
        fdefs += ('<filter id="' + fid + '" x="-50%" y="-50%" width="200%" height="200%">'
                  '<feGaussianBlur stdDeviation="2.5" result="blur"/>'
                  '<feComposite in="SourceGraphic" in2="blur" operator="over"/></filter>')
        ships += ('<g transform="translate(' + str(sx) + ',' + str(sy) + ')">'
            + '<ellipse cx="0" cy="3" rx="13" ry="5" fill="#1D4ED8" opacity="0.85" filter="url(#' + fid + ')"/>'
            + '<rect x="-9" y="-3" width="18" height="7" rx="3" fill="#3B82F6"/>'
            + '<polygon points="9,-3 15,0 9,4" fill="#60A5FA"/>'
            + '<rect x="-4" y="-9" width="3" height="8" fill="#93C5FD" opacity="0.9"/>'
            + '<circle cx="0" cy="0" r="2" fill="#34D399" opacity="0.95"/>'
            + '<text x="0" y="22" font-size="7" fill="#93C5FD" font-weight="700" '
            + 'text-anchor="middle" font-family="Arial,sans-serif">' + lbl + '</text></g>')
    wp0 = WP[0]; wp4 = WP[4]; wp6 = WP[6]
    svg = (
        '<svg viewBox="0 0 800 200" xmlns="http://www.w3.org/2000/svg" '
        'style="width:100%;height:190px;border-radius:12px;display:block;'
        'background:linear-gradient(180deg,#0F172A 0%,#0C2340 40%,#0F3460 100%)">'
        '<defs>'
        '<pattern id="og" width="40" height="40" patternUnits="userSpaceOnUse">'
        '<path d="M 40 0 L 0 0 0 40" fill="none" stroke="rgba(255,255,255,0.04)" stroke-width="0.5"/>'
        '</pattern>'
        '<linearGradient id="rg" x1="0" y1="0" x2="1" y2="0">'
        '<stop offset="0%" stop-color="#4361EE"/>'
        '<stop offset="50%" stop-color="#3B82F6"/>'
        '<stop offset="100%" stop-color="#10B981"/>'
        '</linearGradient>'
        '<filter id="ptg" x="-50%" y="-50%" width="200%" height="200%">'
        '<feGaussianBlur stdDeviation="2.5" result="blur"/>'
        '<feComposite in="SourceGraphic" in2="blur" operator="over"/>'
        '</filter>'
        + fdefs + '</defs>'
        + '<rect width="800" height="200" fill="url(#og)"/>'
        + '<path d="' + path_d + '" fill="none" stroke="rgba(255,255,255,0.1)" stroke-width="2" stroke-dasharray="6,4"/>'
        + '<path d="' + path_d + '" fill="none" stroke="url(#rg)" stroke-width="2.5" stroke-linecap="round"/>'
        + '<circle cx="' + str(wp0[0]) + '" cy="' + str(wp0[1]) + '" r="5" fill="#F59E0B" filter="url(#ptg)"/>'
        + '<circle cx="' + str(wp4[0]) + '" cy="' + str(wp4[1]) + '" r="5" fill="#8B5CF6" filter="url(#ptg)"/>'
        + '<circle cx="' + str(wp6[0]) + '" cy="' + str(wp6[1]) + '" r="6" fill="#10B981" filter="url(#ptg)"/>'
        + '<text x="' + str(wp0[0] - 4) + '" y="' + str(wp0[1] + 16) + '" font-size="8" fill="#F59E0B" '
        + 'font-family="Arial,sans-serif" font-weight="700">DEPART</text>'
        + '<text x="' + str(wp6[0] - 30) + '" y="' + str(wp6[1] + 16) + '" font-size="8" fill="#10B981" '
        + 'font-family="Arial,sans-serif" font-weight="700">GHAZAOUET</text>'
        + '<text x="' + str(wp4[0] - 20) + '" y="' + str(wp4[1] - 12) + '" font-size="7.5" '
        + 'fill="rgba(255,200,80,0.6)" font-family="Arial,sans-serif">ALGECIRAS</text>'
        + ships + '</svg>'
    )
    return svg

def generate_conf_commande(data):
    from openpyxl import load_workbook
    wb = load_workbook("CONF-COMMANDE-TEMPLATE.xlsx")
    ws = wb.active
    ws["A5"]  = data["booking"]
    ws["D5"]  = data["semaine"]
    ws["H5"]  = date.today().strftime("%d/%m/%Y")
    adresse_full = (data["client"] + "\n"
                    + data.get("adresse1", "") + "\n"
                    + data.get("adresse2", "") + "\n"
                    + data.get("ville", "") + " " + data.get("pays", ""))
    ws["A8"]  = adresse_full.strip()
    ws["A10"] = data["licence"]
    ws["C10"] = float(str(data["poids_total_lic"]).replace(",", "").replace(" ", ""))
    ws["A13"] = data["navire"]
    ws["C13"] = data["voyage"]
    ws["F13"] = data["pol"]
    ws["I13"] = data["depart"]
    ws["K13"] = data["eta"]
    ws["A16"] = int(data["nb_cnt"])
    solde_avant = float(str(data.get("solde_avant", 0)).replace(",", "").replace(" ", ""))
    ws["A22"] = solde_avant
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ── LOGIN ──────────────────────────────────────────────────────────────────────
def login_page():
    logo_src  = img_to_b64("logo_eden_food.jpg") or img_to_b64("logo_eden_food.png")
    logo_html = ('<img src="' + logo_src + '" style="height:36px;border-radius:8px;display:block">'
                 if logo_src else
                 '<span style="font-size:22px;font-weight:900;color:#fff">EF</span>')
    st.markdown("""
    <style>
    section[data-testid="stSidebar"]{display:none!important}
    section[data-testid="stSidebarCollapsedControl"]{display:none!important}
    .stApp{background:#000!important}
    .block-container{padding:0!important;max-width:100%!important}
    div[data-testid="stForm"]{
      background:rgba(0,0,0,0.45)!important;backdrop-filter:blur(24px)!important;
      border:1px solid rgba(255,255,255,0.06)!important;border-top:none!important;
      border-radius:0 0 20px 20px!important;padding:4px 28px 28px!important;
      box-shadow:0 30px 80px rgba(0,0,0,0.7)!important;
    }
    div[data-testid="stForm"] input{
      background:rgba(255,255,255,0.05)!important;border:1px solid rgba(255,255,255,0.1)!important;
      color:#fff!important;border-radius:10px!important;height:42px!important;font-size:13px!important;
    }
    div[data-testid="stForm"] input::placeholder{color:rgba(255,255,255,0.28)!important}
    div[data-testid="stForm"] label{color:rgba(255,255,255,0)!important;height:0!important;margin:0!important}
    div[data-testid="stForm"] .stTextInput{margin-bottom:10px}
    div[data-testid="stForm"] .stButton>button{
      background:#fff!important;color:#000!important;font-weight:700!important;
      border:none!important;border-radius:10px!important;height:44px!important;font-size:14px!important;margin-top:6px!important;
    }
    </style>
    """, unsafe_allow_html=True)
    components.html(
        '<!DOCTYPE html><html><head>'
        '<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;700;900&display=swap" rel="stylesheet">'
        '<style>*{margin:0;padding:0;box-sizing:border-box;font-family:Inter,sans-serif}'
        'html,body{height:100%;background:#000;overflow:hidden}'
        '.bg{position:fixed;inset:0;background:linear-gradient(180deg,rgba(168,85,247,.42) 0%,rgba(109,40,217,.48) 38%,#000 100%)}'
        '.glow{position:fixed;top:0;left:50%;transform:translateX(-50%);width:130vh;height:58vh;'
        'border-radius:0 0 50% 50%;background:rgba(167,139,250,.22);filter:blur(80px);'
        'animation:pulse 8s ease-in-out infinite}'
        '@keyframes pulse{0%,100%{opacity:.7}50%{opacity:1}}'
        '.outer{position:relative;z-index:10;display:flex;align-items:center;justify-content:center;height:100%}'
        '.card{position:relative;background:rgba(0,0,0,.48);backdrop-filter:blur(28px);'
        'border-radius:20px 20px 0 0;border:1px solid rgba(255,255,255,.06);border-bottom:none;padding:30px 28px 24px;width:380px}'
        '.beam{position:absolute;top:0;left:-50%;height:2px;width:50%;'
        'background:linear-gradient(90deg,transparent,rgba(255,255,255,.85),transparent);'
        'filter:blur(1.5px);animation:bm 3.5s ease-in-out infinite}'
        '@keyframes bm{0%{left:-50%}65%{left:100%}100%{left:100%}}'
        'h1{text-align:center;font-size:22px;font-weight:800;'
        'background:linear-gradient(180deg,#fff,rgba(255,255,255,.78));'
        '-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:5px}'
        '.sub{text-align:center;font-size:12px;color:rgba(255,255,255,.45);margin-bottom:18px}'
        '.badge{display:inline-flex;align-items:center;gap:7px;background:rgba(16,185,129,.13);'
        'border:1px solid rgba(16,185,129,.28);color:rgba(16,185,129,.95);'
        'padding:5px 14px;border-radius:20px;font-size:10px;font-weight:700}'
        '.dot{width:6px;height:6px;border-radius:50%;background:#10B981;animation:blink 2s ease-in-out infinite}'
        '@keyframes blink{0%,100%{opacity:1}50%{opacity:.25}}'
        '.div{height:1px;margin-top:18px;background:linear-gradient(90deg,transparent,rgba(255,255,255,.08),transparent)}'
        '.logo{width:46px;height:46px;border-radius:50%;border:1px solid rgba(255,255,255,.12);'
        'display:flex;align-items:center;justify-content:center;margin:0 auto 14px}'
        '</style></head><body>'
        '<div class="bg"></div><div class="glow"></div>'
        '<div class="outer"><div class="card">'
        '<div style="position:absolute;inset:-1px;border-radius:20px 20px 0 0;overflow:hidden;pointer-events:none">'
        '<div class="beam"></div></div>'
        '<div style="position:relative;z-index:3;text-align:center">'
        '<div class="logo">' + logo_html + '</div>'
        '<h1>Eden Food</h1>'
        '<p class="sub">Logistics Platform &nbsp;&middot;&nbsp; Acces securise</p>'
        '<div style="display:flex;justify-content:center">'
        '<div class="badge"><div class="dot"></div>Systeme operationnel</div>'
        '</div><div class="div"></div>'
        '</div></div></div>'
        '</body></html>',
        height=268, scrolling=False)
    col1, col2, col3 = st.columns([1, 2.1, 1])
    with col2:
        with st.form("login_form"):
            u   = st.text_input("u", placeholder="Identifiant",  label_visibility="collapsed")
            pwd = st.text_input("p", placeholder="Mot de passe", label_visibility="collapsed", type="password")
            if st.form_submit_button("Connexion  →", use_container_width=True, type="primary"):
                ul = u.strip().lower()
                if ul in USERS and USERS[ul]["password"] == pwd:
                    st.session_state.update(authenticated=True, username=ul, role=USERS[ul]["role"])
                    st.rerun()
                else:
                    st.error("Identifiant ou mot de passe incorrect")

if not st.session_state.authenticated:
    login_page()
    st.stop()

# ── CONSTANTES ─────────────────────────────────────────────────────────────────
POIDS_UNIT  = 18.14
CRTNS       = {"TURBO(COLOMBIA)": 1080, "MOIN(COSTA RICA)": 1200}
KGS_PER_CNT = {"MOIN(COSTA RICA)": 1200 * 18.14, "TURBO(COLOMBIA)": 1080 * 18.14}

def licence_to_filename(lic):
    return str(lic).replace(" ", "_").replace("/", "_") + ".pdf"

def licence_pdf_path(lic):
    return os.path.join("licences", licence_to_filename(lic))

# ── DATA ───────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=60)
def load_clients():
    df = pd.read_excel("eden_food.xlsx", sheet_name="📋 BASE CLIENTS",
                       usecols=list(range(11)), header=3)
    df.columns = ["num", "nom_client", "nom", "adresse", "ville", "pays",
                  "licence", "poids_total", "solde_excel", "cnt_cr", "cnt_col"]
    df = df[df["nom"].notna() & (df["nom"].astype(str).str.strip() != "")].copy()
    df["nom_client"] = df["nom_client"].ffill()
    df["nom_client"] = df["nom_client"].fillna("").astype(str).str.strip()
    return df

@st.cache_data(ttl=60)
def load_commandes():
    df = pd.read_excel("eden_food.xlsx", sheet_name="🚢 COMMANDES",
                       usecols=list(range(13)), header=3)
    df.columns = ["num", "semaine", "client", "booking", "licence", "navire", "voyage",
                  "pol", "depart", "eta", "nb_cnt", "produit", "statut"]
    df = df[df["client"].notna() & (df["client"] != "")].copy()
    df["nb_cnt"]      = pd.to_numeric(df["nb_cnt"], errors="coerce").fillna(0).astype(int)
    df["crtns_cnt"]   = df["pol"].apply(lambda x: CRTNS.get(str(x).strip(), 1200))
    df["total_crtns"] = df["nb_cnt"] * df["crtns_cnt"]
    df["total_kgs"]   = (df["total_crtns"] * POIDS_UNIT).round(2)
    return df

clients_base   = load_clients()
commandes_base = load_commandes()

if st.session_state.new_commandes:
    df_sess = pd.DataFrame(st.session_state.new_commandes)
    df_sess["nb_cnt"]      = pd.to_numeric(df_sess["nb_cnt"], errors="coerce").fillna(0).astype(int)
    df_sess["crtns_cnt"]   = df_sess["pol"].apply(lambda x: CRTNS.get(str(x).strip(), 1200))
    df_sess["total_crtns"] = df_sess["nb_cnt"] * df_sess["crtns_cnt"]
    df_sess["total_kgs"]   = (df_sess["total_crtns"] * POIDS_UNIT).round(2)
    commandes = pd.concat([commandes_base, df_sess], ignore_index=True)
else:
    commandes = commandes_base.copy()

clients = clients_base.copy()
clients["poids_total"] = pd.to_numeric(clients["poids_total"], errors="coerce").fillna(0)

def get_solde_reel(nom, lic, poids):
    mask = ((commandes["client"] == nom) & (commandes["licence"] == lic) &
            (~commandes["statut"].str.contains("A GENERER", na=False)))
    return round(poids - commandes.loc[mask, "total_kgs"].sum(), 2)

def get_solde_prev(nom, lic, poids):
    mask = (commandes["client"] == nom) & (commandes["licence"] == lic)
    return round(poids - commandes.loc[mask, "total_kgs"].sum(), 2)

clients["solde_reel"]   = clients.apply(lambda r: get_solde_reel(r["nom"], r["licence"], r["poids_total"]), axis=1)
clients["solde_prev"]   = clients.apply(lambda r: get_solde_prev(r["nom"], r["licence"], r["poids_total"]), axis=1)
clients["cnt_reel_cr"]  = (clients["solde_reel"] / KGS_PER_CNT["MOIN(COSTA RICA)"]).round(2)
clients["cnt_reel_col"] = (clients["solde_reel"] / KGS_PER_CNT["TURBO(COLOMBIA)"]).round(2)
clients["cnt_prev_cr"]  = (clients["solde_prev"] / KGS_PER_CNT["MOIN(COSTA RICA)"]).round(2)
clients["cnt_prev_col"] = (clients["solde_prev"] / KGS_PER_CNT["TURBO(COLOMBIA)"]).round(2)

current_week_num  = datetime.now().isocalendar()[1]
current_week_str  = "S-" + str(current_week_num)
commandes_semaine = commandes[commandes["semaine"] == current_week_str]

# ── ICONS & NAV ────────────────────────────────────────────────────────────────
ICO = {
    "dashboard": '<svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="3" width="7" height="7"/><rect x="14" y="3" width="7" height="7"/><rect x="14" y="14" width="7" height="7"/><rect x="3" y="14" width="7" height="7"/></svg>',
    "semaine":   '<svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="4" width="18" height="18" rx="2"/><line x1="16" y1="2" x2="16" y2="6"/><line x1="8" y1="2" x2="8" y2="6"/><line x1="3" y1="10" x2="21" y2="10"/></svg>',
    "commandes": '<svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 10c0 7-9 13-9 13s-9-6-9-13a9 9 0 0 1 18 0z"/><circle cx="12" cy="10" r="3"/></svg>',
    "tracking":  '<svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="12" cy="12" r="10"/><polyline points="12 6 12 12 16 14"/></svg>',
    "documents": '<svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/></svg>',
    "licences":  '<svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M9 5H7a2 2 0 0 0-2 2v12a2 2 0 0 0 2 2h10a2 2 0 0 0 2-2V7a2 2 0 0 0-2-2h-2"/><rect x="9" y="3" width="6" height="4" rx="2"/></svg>',
    "planning":  '<svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M20 21v-2a4 4 0 0 0-4-4H8a4 4 0 0 0-4 4v2"/><circle cx="12" cy="7" r="4"/></svg>',
    "new_cmd":   '<svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><line x1="12" y1="5" x2="12" y2="19"/><line x1="5" y1="12" x2="19" y2="12"/></svg>',
    "logout":    '<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"/><polyline points="16 17 21 12 16 7"/><line x1="21" y1="12" x2="9" y2="12"/></svg>',
    "settings":  '<svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="12" cy="12" r="3"/><path d="M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 0 1-2.83 2.83l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-4 0v-.09A1.65 1.65 0 0 0 9 19.4a1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 0 1-2.83-2.83l.06-.06A1.65 1.65 0 0 0 4.68 15a1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1 0-4h.09A1.65 1.65 0 0 0 4.6 9a1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 0 1 2.83-2.83l.06.06A1.65 1.65 0 0 0 9 4.68a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 4 0v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 0 1 2.83 2.83l-.06.06A1.65 1.65 0 0 0 19.4 9a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 0 4h-.09a1.65 1.65 0 0 0-1.51 1z"/></svg>',
}
EMO = {
    "dashboard": "⊞", "semaine": "📅", "commandes": "🚢", "tracking": "🗺️",
    "documents": "📁", "licences": "📋", "planning": "👤",
    "new_cmd": "➕", "logout": "🚪",
}
NAV = [
    ("dashboard", "Overview"),
    ("semaine",   "Semaine " + current_week_str),
    ("commandes", "Commandes"),
    ("tracking",  "Tracking maritime"),
    ("documents", "Documents"),
    ("licences",  "Licences DPVCT"),
    ("planning",  "Planning client"),
]

# ── SIDEBAR ────────────────────────────────────────────────────────────────────
with st.sidebar:
    logo_src = img_to_b64("logo_eden_food.jpg") or img_to_b64("logo_eden_food.png")
    logo_tag = ('<img src="' + logo_src + '" style="height:26px">'
                if logo_src else
                '<span style="font-size:14px;font-weight:800;color:#fff">EDEN FOOD</span>')
    st.markdown(
        '<div style="padding:20px 16px 10px;display:flex;align-items:center;justify-content:space-between">'
        '<div>' + logo_tag + '</div>'
        '<div style="width:7px;height:7px;border-radius:50%;background:#10B981;box-shadow:0 0 7px #10B981"></div>'
        '</div>'
        '<div style="height:1px;background:rgba(255,255,255,0.07);margin:4px 16px 12px"></div>'
        '<div style="padding:0 16px 6px;font-size:9px;color:rgba(255,255,255,0.28);'
        'text-transform:uppercase;letter-spacing:1.8px;font-weight:700">Menu</div>',
        unsafe_allow_html=True)
    for pid, label in NAV:
        is_active = st.session_state.page == pid
        if is_active:
            st.markdown(
                '<div style="background:rgba(255,255,255,0.1);border-radius:8px;'
                'padding:9px 14px;margin-bottom:2px;display:flex;align-items:center;gap:10px;'
                'border-left:2px solid #fff;cursor:default">'
                '<span style="color:#fff;display:flex;align-items:center;flex-shrink:0">' + ICO[pid] + '</span>'
                '<span style="font-size:13px;font-weight:600;color:#fff">' + label + '</span>'
                '</div>',
                unsafe_allow_html=True)
        else:
            if st.button(EMO[pid] + "  " + label, key="nav_" + pid, use_container_width=True):
                st.session_state.page = pid
                st.rerun()
    if st.session_state.role == "admin":
        st.markdown(
            '<div style="height:1px;background:rgba(255,255,255,0.07);margin:8px 16px"></div>'
            '<div style="padding:0 16px 6px;font-size:9px;color:rgba(255,255,255,0.28);'
            'text-transform:uppercase;letter-spacing:1.8px;font-weight:700;margin-top:8px">Admin</div>',
            unsafe_allow_html=True)
        is_new = st.session_state.page == "new_cmd"
        if is_new:
            st.markdown(
                '<div style="background:rgba(255,255,255,0.1);border-radius:8px;'
                'padding:9px 14px;margin-bottom:2px;display:flex;align-items:center;gap:10px;'
                'border-left:2px solid #fff">'
                '<span style="color:#fff;display:flex;align-items:center;flex-shrink:0">' + ICO["new_cmd"] + '</span>'
                '<span style="font-size:13px;font-weight:600;color:#fff">Nouvelle commande</span>'
                '</div>',
                unsafe_allow_html=True)
        else:
            if st.button(EMO["new_cmd"] + "  Nouvelle commande", key="nav_new_cmd", use_container_width=True):
                st.session_state.page = "new_cmd"
                st.rerun()
    st.markdown('<div style="min-height:30px"></div>', unsafe_allow_html=True)
    st.markdown('<div style="height:1px;background:rgba(255,255,255,0.07);margin:0 16px 10px"></div>',
                unsafe_allow_html=True)
    role_color = "#4361EE" if st.session_state.role == "admin" else "#374151"
    role_label = "Admin" if st.session_state.role == "admin" else "Utilisateur"
    initiale   = st.session_state.username[0].upper() if st.session_state.username else "?"
    st.markdown(
        '<div style="padding:10px 12px;display:flex;align-items:center;gap:10px;'
        'background:rgba(255,255,255,0.05);border-radius:10px;margin:0 8px 8px">'
        '<div style="width:34px;height:34px;border-radius:50%;background:' + role_color + ';'
        'display:flex;align-items:center;justify-content:center;font-size:13px;'
        'font-weight:700;color:#fff;flex-shrink:0">' + initiale + '</div>'
        '<div style="min-width:0;flex:1">'
        '<div style="font-size:12px;font-weight:700;color:#fff;white-space:nowrap;'
        'overflow:hidden;text-overflow:ellipsis">' + st.session_state.username.upper() + '</div>'
        '<div style="font-size:10px;color:rgba(255,255,255,0.35)">' + role_label + '</div>'
        '</div>'
        '<span style="color:rgba(255,255,255,0.3);flex-shrink:0">' + ICO["settings"] + '</span>'
        '</div>',
        unsafe_allow_html=True)
    if st.button(EMO["logout"] + "  Deconnexion", use_container_width=True, key="logout"):
        st.session_state.update(authenticated=False, username="")
        st.rerun()

page = st.session_state.page

# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
if page == "dashboard":
    hero_src = img_to_b64("hero.jpg") or img_to_b64("hero.png")
    logo_src = img_to_b64("logo_eden_food.jpg") or img_to_b64("logo_eden_food.png")
    logo_ov  = ('<img src="' + logo_src + '" style="height:36px;margin-bottom:12px;display:block">'
                if logo_src else "")
    if hero_src:
        st.markdown(
            '<div class="hero-wrap" style="background:url(\'' + hero_src + '\') center/cover no-repeat;">'
            '<div class="hero-overlay"><div class="hero-text">'
            + logo_ov
            + '<h1>Fresh from the<br>plantation to the world</h1>'
            '<p>Suivi en temps reel — Colombie & Costa Rica</p>'
            '<div class="hero-badge">🍌 ' + current_week_str + ' · ' + str(len(commandes)) + ' expeditions actives</div>'
            '</div></div></div>',
            unsafe_allow_html=True)
    else:
        st.markdown(
            '<div style="background:linear-gradient(110deg,#0F172A,#4361EE);padding:52px;'
            'height:320px;display:flex;align-items:center;">'
            + logo_ov
            + '<div><h1 style="font-size:32px;font-weight:900;color:#fff;margin:0 0 10px">'
            'Fresh from the plantation to the world</h1>'
            '<p style="font-size:14px;color:rgba(255,255,255,0.72);max-width:440px;margin:0 0 22px">'
            'Colombie & Costa Rica — Suivi logistique temps reel</p>'
            '<div class="hero-badge">🍌 ' + current_week_str + ' · ' + str(len(commandes)) + ' expeditions</div>'
            '</div></div>',
            unsafe_allow_html=True)

    st.markdown('<div class="main-wrap">', unsafe_allow_html=True)
    todo  = commandes[commandes["statut"].str.contains("GENERER", na=False)]
    done  = commandes[commandes["statut"].str.contains("GENERE",  na=False)]
    alert = clients[clients["solde_reel"] < 19591.2]
    tdocs = sum(len(list_docs(b)) for b in commandes["booking"].dropna().unique())

    st.markdown(
        '<div class="kpi-grid">'
        '<div class="kpi-card" style="--card-color:#4361EE;--card-bg:#EEF2FF">'
        '<div class="icon-wrap">🚢</div><div class="kpi-lbl">Expeditions</div>'
        '<div class="kpi-val" style="color:#4361EE">' + str(len(commandes)) + '</div>'
        '<div class="kpi-sub">enregistrees</div></div>'
        '<div class="kpi-card" style="--card-color:#F59E0B;--card-bg:#FEF3C7">'
        '<div class="icon-wrap">⏳</div><div class="kpi-lbl">A generer</div>'
        '<div class="kpi-val" style="color:#D97706">' + str(len(todo)) + '</div>'
        '<div class="kpi-sub">en attente</div></div>'
        '<div class="kpi-card" style="--card-color:#10B981;--card-bg:#D1FAE5">'
        '<div class="icon-wrap">✅</div><div class="kpi-lbl">Confirmees</div>'
        '<div class="kpi-val" style="color:#059669">' + str(len(done)) + '</div>'
        '<div class="kpi-sub">generees</div></div>'
        '<div class="kpi-card" style="--card-color:#8B5CF6;--card-bg:#EDE9FE">'
        '<div class="icon-wrap">📦</div><div class="kpi-lbl">CNT planifies</div>'
        '<div class="kpi-val" style="color:#7C3AED">' + str(int(todo["nb_cnt"].sum())) + '</div>'
        '<div class="kpi-sub">conteneurs</div></div>'
        '<div class="kpi-card" style="--card-color:#EF4444;--card-bg:#FEE2E2">'
        '<div class="icon-wrap">🔴</div><div class="kpi-lbl">Alertes</div>'
        '<div class="kpi-val" style="color:#DC2626">' + str(len(alert)) + '</div>'
        '<div class="kpi-sub">licences critiques</div></div>'
        '<div class="kpi-card" style="--card-color:#3B82F6;--card-bg:#DBEAFE">'
        '<div class="icon-wrap">📁</div><div class="kpi-lbl">Documents</div>'
        '<div class="kpi-val" style="color:#2563EB">' + str(tdocs) + '</div>'
        '<div class="kpi-sub">uploades</div></div>'
        '</div>',
        unsafe_allow_html=True)

    c1, c2 = st.columns([3, 2])
    with c1:
        st.markdown('<div class="sec-hdr"><span class="sec-title">Dernieres expeditions</span>'
                    '<span class="sec-sub">10 dernieres</span></div>', unsafe_allow_html=True)
        st.dataframe(
            commandes.tail(10)[["semaine", "client", "booking", "pol", "nb_cnt", "depart", "statut"]],
            use_container_width=True, hide_index=True, height=340,
            column_config={
                "semaine": st.column_config.TextColumn("Sem."),
                "client":  st.column_config.TextColumn("Client"),
                "booking": st.column_config.TextColumn("Booking"),
                "pol":     st.column_config.TextColumn("POL"),
                "nb_cnt":  st.column_config.NumberColumn("CNT", format="%d"),
                "depart":  st.column_config.TextColumn("Depart"),
                "statut":  st.column_config.TextColumn("Statut"),
            })
    with c2:
        st.markdown('<div class="sec-hdr"><span class="sec-title">Repartition POL</span></div>',
                    unsafe_allow_html=True)
        if not commandes.empty:
            df_pol = commandes.groupby("pol")["nb_cnt"].sum().reset_index()
            fig    = px.pie(df_pol, values="nb_cnt", names="pol", hole=0.65,
                            color_discrete_sequence=["#4361EE", "#F59E0B"])
            fig.update_traces(textinfo="percent+label")
            fig = apply_chart_style(fig, "rgba(0,0,0,0)")
            fig.update_layout(legend=dict(orientation="h", y=-0.1))
            st.plotly_chart(fig, use_container_width=True)

    st.markdown('<div class="sec-hdr"><span class="sec-title">Volume hebdomadaire</span>'
                '<span class="sec-sub">CNT par semaine</span></div>', unsafe_allow_html=True)
    if not commandes.empty:
        df_sem = commandes.groupby("semaine")["nb_cnt"].sum().reset_index()
        fig2   = px.bar(df_sem, x="semaine", y="nb_cnt", color_discrete_sequence=["#4361EE"],
                        labels={"semaine": "", "nb_cnt": "Conteneurs"})
        fig2.update_traces(marker_cornerradius=6, marker_line_width=0)
        fig2 = apply_chart_style(fig2)
        st.plotly_chart(fig2, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# SEMAINE
# ══════════════════════════════════════════════════════════════════════════════
elif page == "semaine":
    st.markdown(
        '<div style="background:#fff;border-bottom:1px solid #E5E7EB;padding:18px 36px;'
        'display:flex;align-items:center;gap:14px;position:sticky;top:0;z-index:100">'
        '<div style="font-size:20px;font-weight:800;color:#111827">Semaine en cours</div>'
        '<div style="background:#EEF2FF;color:#4361EE;padding:4px 12px;border-radius:20px;'
        'font-size:11px;font-weight:700">' + current_week_str + '</div></div>',
        unsafe_allow_html=True)
    st.markdown('<div class="main-wrap">', unsafe_allow_html=True)
    if commandes_semaine.empty:
        st.info("Aucune commande pour " + current_week_str + ".")
    else:
        todo_s = commandes_semaine[commandes_semaine["statut"].str.contains("GENERER", na=False)]
        done_s = commandes_semaine[commandes_semaine["statut"].str.contains("GENERE",  na=False)]
        st.markdown(
            '<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:26px">'
            '<div class="kpi-card" style="--card-color:#4361EE;--card-bg:#EEF2FF">'
            '<div class="kpi-lbl">Total</div>'
            '<div class="kpi-val" style="color:#4361EE">' + str(len(commandes_semaine)) + '</div></div>'
            '<div class="kpi-card" style="--card-color:#F59E0B;--card-bg:#FEF3C7">'
            '<div class="kpi-lbl">A generer</div>'
            '<div class="kpi-val" style="color:#D97706">' + str(len(todo_s)) + '</div></div>'
            '<div class="kpi-card" style="--card-color:#10B981;--card-bg:#D1FAE5">'
            '<div class="kpi-lbl">Confirmees</div>'
            '<div class="kpi-val" style="color:#059669">' + str(len(done_s)) + '</div></div>'
            '<div class="kpi-card" style="--card-color:#8B5CF6;--card-bg:#EDE9FE">'
            '<div class="kpi-lbl">CNT total</div>'
            '<div class="kpi-val" style="color:#7C3AED">' + str(int(commandes_semaine["nb_cnt"].sum())) + '</div></div>'
            '</div>',
            unsafe_allow_html=True)
        for _, row in commandes_semaine.iterrows():
            pc = "pill-green" if "GENERE" in str(row["statut"]) else "pill-orange"
            nd = len(list_docs(row["booking"]))
            dp = '<span class="pill pill-blue">📁 ' + str(nd) + '</span>' if nd else ""
            st.markdown(
                '<div class="cmd-row">'
                '<div style="min-width:190px">'
                '<div style="font-size:14px;font-weight:700;color:#111827">' + str(row["client"]) + '</div>'
                '<div style="font-size:11px;color:#9CA3AF;margin-top:3px">'
                + str(row["booking"]) + ' · ' + str(row["licence"]) + '</div></div>'
                '<div style="text-align:center"><div style="font-size:9px;color:#9CA3AF;margin-bottom:3px;text-transform:uppercase">POL</div>'
                '<div style="font-weight:700;color:#111827;font-size:12px">' + str(row["pol"]) + '</div></div>'
                '<div style="text-align:center"><div style="font-size:9px;color:#9CA3AF;margin-bottom:3px;text-transform:uppercase">CNT</div>'
                '<div style="font-weight:900;color:#4361EE;font-size:22px">' + str(row["nb_cnt"]) + '</div></div>'
                '<div style="text-align:center"><div style="font-size:9px;color:#9CA3AF;margin-bottom:3px;text-transform:uppercase">Depart</div>'
                '<div style="font-weight:600;color:#374151;font-size:12px">' + str(row["depart"]) + '</div></div>'
                '<div style="text-align:center"><div style="font-size:9px;color:#9CA3AF;margin-bottom:3px;text-transform:uppercase">ETA</div>'
                '<div style="font-weight:600;color:#374151;font-size:12px">' + str(row["eta"]) + '</div></div>'
                '<div style="display:flex;gap:6px;align-items:center">'
                + dp + '<span class="pill ' + pc + '">' + str(row["statut"]) + '</span></div>'
                '</div>',
                unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# COMMANDES
# ══════════════════════════════════════════════════════════════════════════════
elif page == "commandes":
    st.markdown(
        '<div style="background:#fff;border-bottom:1px solid #E5E7EB;padding:18px 36px;'
        'position:sticky;top:0;z-index:100;font-size:20px;font-weight:800;color:#111827">'
        'Commandes</div>',
        unsafe_allow_html=True)
    st.markdown('<div style="background:#fff;border-bottom:1px solid #E5E7EB;padding:10px 36px">',
                unsafe_allow_html=True)
    fc1, fc2, fc3, fc4 = st.columns([2, 2, 2, 1])
    with fc1:
        f_client = st.multiselect("c", commandes["client"].dropna().unique().tolist(),
                                  default=commandes["client"].dropna().unique().tolist(),
                                  label_visibility="collapsed", placeholder="Client")
    with fc2:
        f_pol = st.multiselect("p", commandes["pol"].dropna().unique().tolist(),
                               default=commandes["pol"].dropna().unique().tolist(),
                               label_visibility="collapsed", placeholder="POL")
    with fc3:
        f_statut = st.multiselect("s", commandes["statut"].dropna().unique().tolist(),
                                  default=commandes["statut"].dropna().unique().tolist(),
                                  label_visibility="collapsed", placeholder="Statut")
    with fc4:
        f_sem = st.text_input("sem", placeholder="S-18", label_visibility="collapsed")
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('<div class="main-wrap">', unsafe_allow_html=True)
    df_filt = commandes[
        commandes["client"].isin(f_client) &
        commandes["pol"].isin(f_pol) &
        commandes["statut"].isin(f_statut)
    ]
    if f_sem:
        df_filt = df_filt[df_filt["semaine"].str.contains(f_sem, case=False, na=False)]
    st.markdown(
        '<div style="font-size:12px;color:#6B7280;margin-bottom:14px">'
        '<b style="color:#111827">' + str(len(df_filt)) + '</b> commandes · '
        '<b style="color:#111827">' + str(int(df_filt["nb_cnt"].sum())) + '</b> CNT · '
        '<b style="color:#111827">' + "{:,.0f}".format(df_filt["total_kgs"].sum()) + '</b> kgs</div>',
        unsafe_allow_html=True)
    for _, row in df_filt.iterrows():
        pc = "pill-green" if "GENERE" in str(row["statut"]) else "pill-orange"
        nd = len(list_docs(row["booking"]))
        dp = ('<span class="pill pill-blue">📁 ' + str(nd) + '</span>' if nd
              else '<span class="pill" style="background:#F3F4F6;color:#9CA3AF">📁 0</span>')
        st.markdown(
            '<div class="cmd-row">'
            '<div style="min-width:200px">'
            '<div style="font-size:13px;font-weight:700;color:#111827">' + str(row["client"]) + '</div>'
            '<div style="font-size:11px;color:#9CA3AF;margin-top:2px">'
            + str(row["semaine"]) + ' · ' + str(row["booking"]) + '</div></div>'
            '<div style="text-align:center"><div style="font-size:9px;color:#9CA3AF;margin-bottom:2px;text-transform:uppercase">Navire</div>'
            '<div style="font-weight:500;color:#374151;font-size:11px">' + str(row["navire"]) + '</div></div>'
            '<div style="text-align:center"><div style="font-size:9px;color:#9CA3AF;margin-bottom:2px;text-transform:uppercase">POL</div>'
            '<div style="font-weight:700;color:#111827;font-size:12px">' + str(row["pol"]) + '</div></div>'
            '<div style="text-align:center"><div style="font-size:9px;color:#9CA3AF;margin-bottom:2px;text-transform:uppercase">CNT</div>'
            '<div style="font-weight:900;color:#4361EE;font-size:20px">' + str(row["nb_cnt"]) + '</div></div>'
            '<div style="text-align:center"><div style="font-size:9px;color:#9CA3AF;margin-bottom:2px;text-transform:uppercase">Depart</div>'
            '<div style="font-weight:500;color:#374151;font-size:11px">' + str(row["depart"]) + '</div></div>'
            '<div style="text-align:center"><div style="font-size:9px;color:#9CA3AF;margin-bottom:2px;text-transform:uppercase">ETA</div>'
            '<div style="font-weight:500;color:#374151;font-size:11px">' + str(row["eta"]) + '</div></div>'
            '<div style="display:flex;gap:6px;align-items:center">'
            + dp + '<span class="pill ' + pc + '">' + str(row["statut"]) + '</span></div>'
            '</div>',
            unsafe_allow_html=True)
        is_open = st.session_state.expanded_cmd == row["booking"]
        col_btn, _ = st.columns([1, 5])
        with col_btn:
            if st.button("Fermer" if is_open else "📁 Documents",
                         key="docbtn_" + str(row["booking"]), use_container_width=True):
                st.session_state.expanded_cmd = None if is_open else row["booking"]
                st.rerun()
        if is_open:
            existing = list_docs(row["booking"])
            if existing:
                for doc_name in existing:
                    doc_path = os.path.join(docs_path(row["booking"]), doc_name)
                    d1, d2, d3 = st.columns([4, 1, 1])
                    with d1:
                        st.markdown('<div class="doc-chip">📄 ' + doc_name + '</div>', unsafe_allow_html=True)
                    with d2:
                        with open(doc_path, "rb") as f:
                            st.download_button("⬇️", f.read(), file_name=doc_name,
                                key="dl_" + str(row["booking"]) + "_" + doc_name, use_container_width=True)
                    with d3:
                        if st.session_state.role == "admin":
                            if st.button("🗑️", key="del_" + str(row["booking"]) + "_" + doc_name,
                                         use_container_width=True):
                                delete_doc(row["booking"], doc_name)
                                st.rerun()
            else:
                st.caption("Aucun document.")
            u1, u2 = st.columns([3, 1])
            with u1:
                uploaded = st.file_uploader("", type=["pdf", "xlsx", "xls", "docx", "jpg", "png"],
                                            key="up_" + str(row["booking"]), label_visibility="collapsed")
            with u2:
                dtype = st.selectbox("", DOC_TYPES, key="dtype_" + str(row["booking"]), label_visibility="collapsed")
            if uploaded and st.button("✅ Uploader", key="upconf_" + str(row["booking"]), type="primary"):
                uploaded.name = dtype.replace(" ", "_") + "_" + uploaded.name
                save_doc(row["booking"], uploaded)
                st.success("Uploade !")
                st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# TRACKING
# ══════════════════════════════════════════════════════════════════════════════
elif page == "tracking":
    st.markdown(
        '<div style="background:#fff;border-bottom:1px solid #E5E7EB;padding:18px 36px;'
        'display:flex;align-items:center;gap:14px;position:sticky;top:0;z-index:100">'
        '<div style="font-size:20px;font-weight:800;color:#111827">Tracking maritime</div>'
        '<div style="background:#DBEAFE;color:#1D4ED8;padding:4px 12px;border-radius:20px;'
        'font-size:11px;font-weight:700">En temps reel</div></div>',
        unsafe_allow_html=True)
    st.markdown('<div class="main-wrap">', unsafe_allow_html=True)
    df_track = commandes[commandes["depart"].notna()].copy()
    en_attente_list, en_mer_list, arrives_list = [], [], []
    active_ships_for_map = []
    for _, row in df_track.iterrows():
        s_idx, prog, days_el, eta_calc = get_tracking_info(row["depart"], row["pol"])
        routing    = get_routing(row["pol"])
        total_days = routing["total_days"]
        if days_el is not None and days_el < 0:
            en_attente_list.append(row)
        elif days_el is not None and days_el >= total_days:
            arrives_list.append(row)
        else:
            en_mer_list.append(row)
            active_ships_for_map.append({"progress": prog, "label": str(row["client"])[:10]})
    st.markdown(
        '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:24px">'
        '<div class="kpi-card" style="--card-color:#F59E0B;--card-bg:#FEF3C7">'
        '<div class="icon-wrap">⏳</div><div class="kpi-lbl">En attente depart</div>'
        '<div class="kpi-val" style="color:#D97706">' + str(len(en_attente_list)) + '</div></div>'
        '<div class="kpi-card" style="--card-color:#3B82F6;--card-bg:#DBEAFE">'
        '<div class="icon-wrap">🚢</div><div class="kpi-lbl">En mer</div>'
        '<div class="kpi-val" style="color:#2563EB">' + str(len(en_mer_list)) + '</div></div>'
        '<div class="kpi-card" style="--card-color:#10B981;--card-bg:#D1FAE5">'
        '<div class="icon-wrap">⚓</div><div class="kpi-lbl">Arrives Ghazaouet</div>'
        '<div class="kpi-val" style="color:#059669">' + str(len(arrives_list)) + '</div></div>'
        '</div>',
        unsafe_allow_html=True)
    map_svg = render_route_map_svg(active_ships_for_map)
    components.html(
        '<div style="background:#0F172A;border-radius:16px;padding:16px;'
        'box-shadow:0 8px 32px rgba(0,0,0,0.3);margin-bottom:20px">'
        '<div style="font-size:11px;color:rgba(255,255,255,0.5);font-weight:700;'
        'text-transform:uppercase;letter-spacing:1.5px;margin-bottom:10px;'
        'font-family:Arial,sans-serif">Route maritime — Ameriques vers Ghazaouet</div>'
        + map_svg + '</div>', height=250)
    filter_track = st.selectbox("", ["Tous", "En attente", "En mer", "Arrives"],
                                label_visibility="collapsed")
    if   filter_track == "En attente": rows_to_show = en_attente_list
    elif filter_track == "En mer":     rows_to_show = en_mer_list
    elif filter_track == "Arrives":    rows_to_show = arrives_list
    else:                              rows_to_show = en_mer_list + en_attente_list + arrives_list
    if not rows_to_show:
        st.info("Aucune commande dans cette categorie.")
    else:
        for row in rows_to_show:
            s_idx, prog, days_el, eta_calc = get_tracking_info(row["depart"], row["pol"])
            routing    = get_routing(row["pol"])
            total_days = routing["total_days"]
            stages     = routing["stages"]
            current_s  = stages[min(s_idx, len(stages) - 1)]
            if days_el is not None and days_el >= total_days:
                card_color = "#10B981"; card_bg = "#D1FAE5"; status_txt = "Arrive a Ghazaouet"
            elif s_idx >= 3:
                card_color = "#8B5CF6"; card_bg = "#EDE9FE"; status_txt = "Transit " + current_s["label"]
            elif s_idx >= 1:
                card_color = "#3B82F6"; card_bg = "#DBEAFE"; status_txt = "En mer — " + current_s["label"]
            elif days_el is not None and days_el >= 0:
                card_color = "#4361EE"; card_bg = "#EEF2FF"; status_txt = "Port de depart"
            else:
                card_color = "#F59E0B"; card_bg = "#FEF3C7"; status_txt = "Avant depart"
            if days_el is not None and days_el >= 0:
                days_label = "J+" + str(days_el)
            elif days_el is not None:
                days_label = "Depart dans " + str(abs(days_el)) + "j"
            else:
                days_label = "—"
            eta_label     = eta_calc.strftime("%d/%m/%Y") if eta_calc else "—"
            timeline_html = build_stage_timeline_html(s_idx, prog, str(row["pol"]))
            bkey          = str(row["booking"]).replace(" ", "_").replace("/", "_")
            is_open       = bkey in st.session_state.expanded_track
            st.markdown(
                '<div class="track-card" style="border-top:3px solid ' + card_color + '">'
                '<div class="track-header">'
                '<div>'
                '<div style="font-size:15px;font-weight:800;color:#111827">' + str(row["client"]) + '</div>'
                '<div style="font-size:11px;color:#9CA3AF;margin-top:3px">'
                + str(row["navire"]) + ' · ' + str(row["booking"]) + ' · ' + str(row["pol"]).split("(")[0].strip() + '</div></div>'
                '<div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap">'
                '<div style="text-align:center"><div style="font-size:9px;color:#9CA3AF;text-transform:uppercase;margin-bottom:2px">Depart</div>'
                '<div style="font-size:12px;font-weight:700;color:#374151">' + str(row["depart"]) + '</div></div>'
                '<div style="text-align:center"><div style="font-size:9px;color:#9CA3AF;text-transform:uppercase;margin-bottom:2px">ETA Ghaz.</div>'
                '<div style="font-size:12px;font-weight:700;color:#374151">' + eta_label + '</div></div>'
                '<div style="text-align:center"><div style="font-size:9px;color:#9CA3AF;text-transform:uppercase;margin-bottom:2px">Transit</div>'
                '<div style="font-size:16px;font-weight:900;color:' + card_color + '">' + str(total_days) + 'j</div></div>'
                '<div style="text-align:center"><div style="font-size:9px;color:#9CA3AF;text-transform:uppercase;margin-bottom:2px">Progression</div>'
                '<div style="font-size:16px;font-weight:900;color:' + card_color + '">' + str(int(prog)) + '%</div></div>'
                '<span class="pill" style="background:' + card_bg + ';color:' + card_color + '">' + status_txt + '</span>'
                '<span style="font-size:11px;font-weight:700;color:' + card_color + ';background:' + card_bg + ';padding:4px 10px;border-radius:20px">' + days_label + '</span>'
                '</div></div>'
                '<div style="padding:0 22px 10px">'
                '<div class="prog-track" style="height:6px">'
                '<div class="prog-fill" style="background:' + card_color + ';width:' + str(prog) + '%"></div>'
                '</div></div></div>',
                unsafe_allow_html=True)
            if st.button("Masquer etapes" if is_open else "Voir etapes", key="tt_" + bkey):
                if is_open:
                    st.session_state.expanded_track.discard(bkey)
                else:
                    st.session_state.expanded_track.add(bkey)
                st.rerun()
            if is_open:
                st.markdown(
                    '<div style="background:#F9FAFB;border:1px solid #E5E7EB;border-radius:12px;'
                    'padding:18px 22px;margin-bottom:10px;overflow-x:auto">'
                    + timeline_html
                    + '<div style="margin-top:16px;display:grid;grid-template-columns:repeat(4,1fr);gap:10px">'
                    '<div style="background:#fff;border-radius:10px;padding:12px;border:1px solid #E5E7EB">'
                    '<div style="font-size:9px;color:#9CA3AF;text-transform:uppercase;font-weight:700;margin-bottom:4px">CNT</div>'
                    '<div style="font-size:18px;font-weight:900;color:#4361EE">' + str(row["nb_cnt"]) + '</div></div>'
                    '<div style="background:#fff;border-radius:10px;padding:12px;border:1px solid #E5E7EB">'
                    '<div style="font-size:9px;color:#9CA3AF;text-transform:uppercase;font-weight:700;margin-bottom:4px">Total kgs</div>'
                    '<div style="font-size:14px;font-weight:800;color:#111827">' + "{:,.0f}".format(row["total_kgs"]) + '</div></div>'
                    '<div style="background:#fff;border-radius:10px;padding:12px;border:1px solid #E5E7EB">'
                    '<div style="font-size:9px;color:#9CA3AF;text-transform:uppercase;font-weight:700;margin-bottom:4px">Semaine</div>'
                    '<div style="font-size:14px;font-weight:800;color:#111827">' + str(row["semaine"]) + '</div></div>'
                    '<div style="background:#fff;border-radius:10px;padding:12px;border:1px solid #E5E7EB">'
                    '<div style="font-size:9px;color:#9CA3AF;text-transform:uppercase;font-weight:700;margin-bottom:4px">Statut</div>'
                    '<div style="font-size:11px;font-weight:700;color:' + card_color + '">' + str(row["statut"]) + '</div></div>'
                    '</div></div>',
                    unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# DOCUMENTS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "documents":
    st.markdown(
        '<div style="background:#fff;border-bottom:1px solid #E5E7EB;padding:18px 36px;'
        'position:sticky;top:0;z-index:100;font-size:20px;font-weight:800;color:#111827">'
        'Documents</div>',
        unsafe_allow_html=True)
    st.markdown('<div class="main-wrap">', unsafe_allow_html=True)
    all_bookings = commandes["booking"].dropna().unique().tolist()
    total_docs   = sum(len(list_docs(b)) for b in all_bookings)
    bookings_ok  = sum(1 for b in all_bookings if len(list_docs(b)) > 0)
    st.markdown(
         '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:24px">'
        '<div class="kpi-card" style="--card-color:#4361EE;--card-bg:#EEF2FF">'
        '<div class="icon-wrap">\U0001f4e6</div><div class="kpi-lbl">Bookings total</div>'
        '<div class="kpi-val" style="color:#4361EE">' + str(len(all_bookings)) + '</div></div>'
        '<div class="kpi-card" style="--card-color:#10B981;--card-bg:#D1FAE5">'
        '<div class="icon-wrap">\U0001f4c1</div><div class="kpi-lbl">Avec documents</div>'
        '<div class="kpi-val" style="color:#059669">' + str(bookings_ok) + '</div></div>'
        '<div class="kpi-card" style="--card-color:#8B5CF6;--card-bg:#EDE9FE">'
        '<div class="icon-wrap">\U0001f4c4</div><div class="kpi-lbl">Total fichiers</div>'
        '<div class="kpi-val" style="color:#7C3AED">' + str(total_docs) + '</div></div>'
        '</div>',
        unsafe_allow_html=True)
    search_doc = st.text_input("", placeholder="🔍 Rechercher un booking, client...", label_visibility="collapsed")
    for _, row in commandes.iterrows():
        bk   = str(row["booking"])
        docs = list_docs(bk)
        if search_doc and search_doc.lower() not in bk.lower() and search_doc.lower() not in str(row["client"]).lower():
            continue
        nb_docs = len(docs)
        zone_color = "#D1FAE5" if nb_docs >= 3 else ("#FEF3C7" if nb_docs >= 1 else "#FEE2E2")
        zone_icon  = "✅" if nb_docs >= 3 else ("⚠️" if nb_docs >= 1 else "❌")
        with st.expander(zone_icon + "  " + str(row["client"]) + "  ·  " + bk + "  ·  " + str(row["semaine"]) + "  ·  " + str(nb_docs) + " doc(s)"):
            st.markdown('<div class="doc-zone">', unsafe_allow_html=True)
            if docs:
                for doc_name in docs:
                    doc_path = os.path.join(docs_path(bk), doc_name)
                    d1, d2, d3 = st.columns([4, 1, 1])
                    with d1:
                        st.markdown('<div class="doc-chip">📄 ' + doc_name + '</div>', unsafe_allow_html=True)
                    with d2:
                        with open(doc_path, "rb") as f:
                            st.download_button("⬇️", f.read(), file_name=doc_name,
                                key="dldoc_" + bk + "_" + doc_name, use_container_width=True)
                    with d3:
                        if st.session_state.role == "admin":
                            if st.button("🗑️", key="deldoc_" + bk + "_" + doc_name, use_container_width=True):
                                delete_doc(bk, doc_name)
                                st.rerun()
            else:
                st.caption("Aucun document pour ce booking.")
            st.markdown('</div>', unsafe_allow_html=True)
            u1, u2 = st.columns([3, 1])
            with u1:
                uploaded = st.file_uploader("", type=["pdf", "xlsx", "xls", "docx", "jpg", "png"],
                                            key="updoc_" + bk, label_visibility="collapsed")
            with u2:
                dtype = st.selectbox("", DOC_TYPES, key="dtypedoc_" + bk, label_visibility="collapsed")
            if uploaded and st.button("✅ Uploader", key="upconfdoc_" + bk, type="primary"):
                uploaded.name = dtype.replace(" ", "_") + "_" + uploaded.name
                save_doc(bk, uploaded)
                st.success("Uploade !")
                st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# LICENCES
# ══════════════════════════════════════════════════════════════════════════════
elif page == "licences":
    st.markdown(
        '<div style="background:#fff;border-bottom:1px solid #E5E7EB;padding:18px 36px;'
        'position:sticky;top:0;z-index:100;font-size:20px;font-weight:800;color:#111827">'
        'Licences DPVCT</div>',
        unsafe_allow_html=True)
    st.markdown('<div class="main-wrap">', unsafe_allow_html=True)
    total_poids = clients["poids_total"].sum()
    total_reel  = clients["solde_reel"].sum()
    total_prev  = clients["solde_prev"].sum()
    pct_cons    = round((1 - total_reel / total_poids) * 100, 1) if total_poids else 0
    st.markdown(
        '<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:24px">'
        '<div class="kpi-card" style="--card-color:#4361EE;--card-bg:#EEF2FF">'
        '<div class="icon-wrap">📋</div><div class="kpi-lbl">Total licences</div>'
        '<div class="kpi-val" style="color:#4361EE">' + str(len(clients)) + '</div></div>'
        '<div class="kpi-card" style="--card-color:#10B981;--card-bg:#D1FAE5">'
        '<div class="icon-wrap">⚖️</div><div class="kpi-lbl">Solde reel total</div>'
        '<div class="kpi-val" style="color:#059669;font-size:1.3rem">' + "{:,.0f}".format(total_reel) + '</div>'
        '<div class="kpi-sub">kgs</div></div>'
        '<div class="kpi-card" style="--card-color:#8B5CF6;--card-bg:#EDE9FE">'
        '<div class="icon-wrap">📊</div><div class="kpi-lbl">Consomme</div>'
        '<div class="kpi-val" style="color:#7C3AED">' + str(pct_cons) + '%</div></div>'
        '<div class="kpi-card" style="--card-color:#F59E0B;--card-bg:#FEF3C7">'
        '<div class="icon-wrap">⚠️</div><div class="kpi-lbl">Licences critiques</div>'
        '<div class="kpi-val" style="color:#D97706">'
        + str(len(clients[clients["solde_reel"] < 19591.2])) + '</div></div>'
        '</div>',
        unsafe_allow_html=True)

    search_lic = st.text_input("", placeholder="🔍 Rechercher client, licence...", label_visibility="collapsed")
        for idx, (_, row) in enumerate(clients.iterrows()):
        if search_lic and search_lic.lower() not in str(row["nom"]).lower() and search_lic.lower() not in str(row["licence"]).lower():
            continue
        poids   = row["poids_total"]
        solde   = row["solde_reel"]
        pct     = max(0, min(100, solde / poids * 100)) if poids > 0 else 0
        if pct > 30:
            bar_color = "#10B981"; alert_msg = ""
        elif pct > 10:
            bar_color = "#F59E0B"
            alert_msg = '<div class="alert alert-warn">Solde < 30% — Renouvellement recommande</div>'
        else:
            bar_color = "#EF4444"
            alert_msg = '<div class="alert alert-red">Solde critique — Action urgente requise</div>'

        has_pdf  = os.path.exists(licence_pdf_path(row["licence"]))
        pdf_badge = ('<span class="pill pill-green">PDF OK</span>' if has_pdf
                     else '<span class="pill" style="background:#F3F4F6;color:#9CA3AF">PDF Manquant</span>')
        cnt_cr  = row["cnt_reel_cr"]
        cnt_col = row["cnt_reel_col"]
        uid     = str(idx) + "_" + str(row["licence"]).replace(" ", "_").replace("/", "_")

        with st.expander(str(row["nom"]) + "  ·  " + str(row["licence"]) + "  ·  " + str(int(pct)) + "% restant"):
            st.markdown(alert_msg, unsafe_allow_html=True)
            st.markdown(
                '<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin:12px 0">'
                '<div style="background:#F9FAFB;border-radius:10px;padding:12px;border:1px solid #E5E7EB">'
                '<div style="font-size:9px;color:#9CA3AF;font-weight:700;text-transform:uppercase;margin-bottom:3px">Poids total</div>'
                '<div style="font-size:14px;font-weight:800;color:#111827">' + "{:,.0f}".format(poids) + ' kgs</div></div>'
                '<div style="background:#F9FAFB;border-radius:10px;padding:12px;border:1px solid #E5E7EB">'
                '<div style="font-size:9px;color:#9CA3AF;font-weight:700;text-transform:uppercase;margin-bottom:3px">Solde reel</div>'
                '<div style="font-size:14px;font-weight:800;color:' + bar_color + '">' + "{:,.0f}".format(solde) + ' kgs</div></div>'
                '<div style="background:#F9FAFB;border-radius:10px;padding:12px;border:1px solid #E5E7EB">'
                '<div style="font-size:9px;color:#9CA3AF;font-weight:700;text-transform:uppercase;margin-bottom:3px">CNT dispo CR</div>'
                '<div style="font-size:14px;font-weight:800;color:#4361EE">' + str(cnt_cr) + '</div></div>'
                '<div style="background:#F9FAFB;border-radius:10px;padding:12px;border:1px solid #E5E7EB">'
                '<div style="font-size:9px;color:#9CA3AF;font-weight:700;text-transform:uppercase;margin-bottom:3px">CNT dispo COL</div>'
                '<div style="font-size:14px;font-weight:800;color:#4361EE">' + str(cnt_col) + '</div></div>'
                '</div>',
                unsafe_allow_html=True)
            st.markdown(
                '<div class="prog-track"><div class="prog-fill" style="background:' + bar_color
                + ';width:' + str(pct) + '%"></div></div>',
                unsafe_allow_html=True)
            l1, l2 = st.columns([1, 1])
            with l1:
                st.markdown(pdf_badge, unsafe_allow_html=True)
                if has_pdf:
                    with open(licence_pdf_path(row["licence"]), "rb") as f:
                        st.download_button("Telecharger PDF",
                            f.read(), file_name=licence_to_filename(row["licence"]),
                            key="dllic_" + uid, use_container_width=True)
            with l2:
                if st.session_state.role == "admin":
                    upl = st.file_uploader("", type=["pdf"], key="upllic_" + uid,
                                           label_visibility="collapsed")
                    if upl and st.button("Uploader PDF", key="uplconflic_" + uid, type="primary"):
                        os.makedirs("licences", exist_ok=True)
                        with open(licence_pdf_path(row["licence"]), "wb") as f:
                            f.write(upl.getbuffer())
                        st.success("PDF uploade !")
                        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)
# ══════════════════════════════════════════════════════════════════════════════
# PLANNING CLIENT
# ══════════════════════════════════════════════════════════════════════════════
elif page == "planning":
    st.markdown(
        '<div style="background:#fff;border-bottom:1px solid #E5E7EB;padding:18px 36px;'
        'position:sticky;top:0;z-index:100;font-size:20px;font-weight:800;color:#111827">'
        'Planning client</div>',
        unsafe_allow_html=True)
    st.markdown('<div class="main-wrap">', unsafe_allow_html=True)
    noms_clients = sorted(clients["nom_client"].dropna().str.strip().replace("", pd.NA).dropna().unique().tolist())
    cli_sel = st.selectbox("", ["— Selectionner un client —"] + noms_clients, label_visibility="collapsed")
    if cli_sel == "— Selectionner un client —":
        st.info("Selectionnez un client pour voir son planning.")
    else:
        cli_rows = clients[clients["nom_client"].str.strip() == cli_sel]
        cli_cmds = commandes[commandes["client"].isin(cli_rows["nom"].tolist())]
        nb_soc   = len(cli_rows)
        st.markdown(
            '<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:20px">'
            '<div class="kpi-card" style="--card-color:#4361EE;--card-bg:#EEF2FF">'
            '<div class="kpi-lbl">Licences</div>'
            '<div class="kpi-val" style="color:#4361EE">' + str(nb_soc) + '</div></div>'
            '<div class="kpi-card" style="--card-color:#10B981;--card-bg:#D1FAE5">'
            '<div class="kpi-lbl">Commandes</div>'
            '<div class="kpi-val" style="color:#059669">' + str(len(cli_cmds)) + '</div></div>'
            '<div class="kpi-card" style="--card-color:#8B5CF6;--card-bg:#EDE9FE">'
            '<div class="kpi-lbl">CNT total</div>'
            '<div class="kpi-val" style="color:#7C3AED">' + str(int(cli_cmds["nb_cnt"].sum())) + '</div></div>'
            '<div class="kpi-card" style="--card-color:#F59E0B;--card-bg:#FEF3C7">'
            '<div class="kpi-lbl">Solde reel total</div>'
            '<div class="kpi-val" style="color:#D97706;font-size:1.1rem">'
            + "{:,.0f}".format(cli_rows["solde_reel"].sum()) + '</div></div>'
            '</div>',
            unsafe_allow_html=True)
        st.markdown(
            '<div class="sec-hdr">'
            '<span class="sec-title">Societes du client</span>'
            '<span class="sec-sub">' + str(nb_soc) + ' societe(s)</span>'
            '</div>',
            unsafe_allow_html=True)
        for _, srow in cli_rows.iterrows():
            sr = max(0, min(100, srow["solde_reel"] / srow["poids_total"] * 100)) if srow["poids_total"] > 0 else 0
            sc = "#10B981" if sr > 30 else ("#F59E0B" if sr > 10 else "#EF4444")
            st.markdown(
                '<div class="cmd-row">'
                '<div style="flex:1">'
                '<div style="font-size:13px;font-weight:700;color:#111827">' + str(srow["nom"]) + '</div>'
                '<div style="font-size:11px;color:#9CA3AF;margin-top:2px">'
                + str(srow.get("ville", "")) + ' · Licence ' + str(srow["licence"]) + '</div></div>'
                '<div style="text-align:center;min-width:100px">'
                '<div style="font-size:9px;color:#9CA3AF;text-transform:uppercase;margin-bottom:2px">Solde reel</div>'
                '<div style="font-size:13px;font-weight:800;color:#4361EE">'
                + "{:,.0f}".format(srow["solde_reel"]) + ' kgs</div></div>'
                '<div style="text-align:center;min-width:80px">'
                '<div style="font-size:9px;color:#9CA3AF;text-transform:uppercase;margin-bottom:2px">Reste</div>'
                '<div style="font-size:13px;font-weight:800;color:' + sc + '">' + str(int(sr)) + '%</div></div>'
                '</div>',
                unsafe_allow_html=True)
        st.markdown(
            '<div class="sec-hdr" style="margin-top:20px">'
            '<span class="sec-title">Historique commandes</span>'
            '<span class="sec-sub">' + str(len(cli_cmds)) + ' commandes</span>'
            '</div>',
            unsafe_allow_html=True)
        if cli_cmds.empty:
            st.info("Aucune commande pour ce client.")
        else:
            for _, crow in cli_cmds.iterrows():
                cpill = "pill-green" if "GENERE" in str(crow["statut"]) else "pill-orange"
                st.markdown(
                    '<div class="cmd-row">'
                    '<div style="min-width:120px">'
                    '<div style="font-size:12px;font-weight:700;color:#4361EE">' + str(crow["semaine"]) + '</div>'
                    '<div style="font-size:11px;color:#9CA3AF;margin-top:2px">' + str(crow["booking"]) + '</div>'
                    '</div>'
                    '<div style="text-align:center">'
                    '<div style="font-size:9px;color:#9CA3AF;text-transform:uppercase;margin-bottom:2px">POL</div>'
                    '<div style="font-size:11px;font-weight:700;color:#111827">' + str(crow["pol"]) + '</div>'
                    '</div>'
                    '<div style="text-align:center">'
                    '<div style="font-size:9px;color:#9CA3AF;text-transform:uppercase;margin-bottom:2px">CNT</div>'
                    '<div style="font-size:20px;font-weight:900;color:#4361EE">' + str(crow["nb_cnt"]) + '</div>'
                    '</div>'
                    '<div style="text-align:center">'
                    '<div style="font-size:9px;color:#9CA3AF;text-transform:uppercase;margin-bottom:2px">Depart</div>'
                    '<div style="font-size:11px;font-weight:600;color:#374151">' + str(crow["depart"]) + '</div>'
                    '</div>'
                    '<span class="pill ' + cpill + '">' + str(crow["statut"]) + '</span>'
                    '</div>',
                    unsafe_allow_html=True)
            df_cli_sem = cli_cmds.groupby("semaine")["nb_cnt"].sum().reset_index()
            fig3 = px.bar(df_cli_sem, x="semaine", y="nb_cnt",
                          color_discrete_sequence=["#4361EE"],
                          labels={"semaine": "", "nb_cnt": "CNT"})
            fig3.update_traces(marker_cornerradius=6, marker_line_width=0)
            fig3 = apply_chart_style(fig3)
            st.plotly_chart(fig3, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# NOUVELLE COMMANDE
# ══════════════════════════════════════════════════════════════════════════════
elif page == "new_cmd":
    if st.session_state.role != "admin":
        st.error("Acces refuse.")
        st.stop()
    st.markdown(
        '<div style="background:#fff;border-bottom:1px solid #E5E7EB;padding:18px 36px;'
        'position:sticky;top:0;z-index:100;font-size:20px;font-weight:800;color:#111827">'
        'Nouvelle commande</div>',
        unsafe_allow_html=True)
    st.markdown('<div class="main-wrap">', unsafe_allow_html=True)
    noms_cmd = ["— Selectionner —"] + sorted(
        clients["nom_client"].dropna().str.strip().replace("", pd.NA).dropna().unique().tolist()
    )
    cli_cmd_sel = st.selectbox("Client", noms_cmd, label_visibility="collapsed")
    if cli_cmd_sel != "— Selectionner —":
        cli_cmd_rows   = clients[clients["nom_client"].str.strip() == cli_cmd_sel]
        licences_dispo = cli_cmd_rows[["nom", "licence", "solde_reel"]].copy()
        licences_dispo["label"] = (
            licences_dispo["nom"] + " — " +
            licences_dispo["licence"] + " — Solde: " +
            licences_dispo["solde_reel"].apply(lambda x: "{:,.0f}".format(x)) + " kgs"
        )
        lic_opts = ["— Choisir une licence —"] + licences_dispo["label"].tolist()
        lic_sel  = st.selectbox("Licence", lic_opts, label_visibility="collapsed")
        if lic_sel != "— Choisir une licence —":
            lic_row    = licences_dispo[licences_dispo["label"] == lic_sel].iloc[0]
            client_nom = lic_row["nom"]
            match      = clients[clients["nom"] == client_nom]
            client_row = match.iloc[0] if len(match) > 0 else None
            with st.form("new_cmd_form"):
                c1, c2, c3 = st.columns(3)
                with c1:
                    semaine = st.text_input("Semaine", value=current_week_str)
                    booking = st.text_input("Reference Booking")
                    navire  = st.text_input("Navire")
                with c2:
                    voyage  = st.text_input("N Voyage")
                    pol     = st.selectbox("POL", ["TURBO(COLOMBIA)", "MOIN(COSTA RICA)"])
                    nb_cnt  = st.number_input("Nombre CNT", min_value=1, max_value=500, value=80)
                with c3:
                    depart_d = st.date_input("Date depart", value=date.today())
                    st.number_input("Solde avant (kgs)", value=float(lic_row["solde_reel"]), format="%.2f")
                    produit  = st.text_input("Produit", value="BANANE")
                if st.form_submit_button("Creer la commande", type="primary", use_container_width=True):
                    routing_sel = get_routing(pol)
                    eta_calc    = depart_d + timedelta(days=routing_sel["total_days"])
                    st.session_state.new_commandes.append({
                        "num": "", "semaine": semaine, "client": client_nom,
                        "booking": booking, "licence": lic_row["licence"],
                        "navire": navire, "voyage": voyage, "pol": pol,
                        "depart": depart_d.strftime("%d/%m/%Y"),
                        "eta":    eta_calc.strftime("%d/%m/%Y"),
                        "nb_cnt": nb_cnt, "produit": produit,
                        "statut": "A GENERER",
                        "total_kgs": nb_cnt * CRTNS.get(pol, 1200) * POIDS_UNIT,
                    })
                    st.success("Commande creee ! Allez dans Commandes pour la voir.")
                    st.rerun()
            adresse1  = str(client_row.get("adresse", "")) if client_row is not None else ""
            ville     = str(client_row.get("ville", ""))   if client_row is not None else ""
            pays      = str(client_row.get("pays", ""))    if client_row is not None else ""
            poids_lic = client_row.get("poids_total", 0)   if client_row is not None else 0
            routing_prev = get_routing(pol if 'pol' in dir() else "MOIN(COSTA RICA)")
            eta_prev = (date.today() + timedelta(days=routing_prev["total_days"])).strftime("%d/%m/%Y")
            if st.button("📄 Generer confirmation commande Excel", type="primary", use_container_width=True,
                         key="gen_conf"):
                try:
                    xlsx_bytes = generate_conf_commande({
                        "booking":       booking if 'booking' in dir() else "",
                        "semaine":       semaine if 'semaine' in dir() else current_week_str,
                        "client":        client_nom,
                        "adresse1":      adresse1,
                        "adresse2":      "",
                        "ville":         ville,
                        "pays":          pays,
                        "licence":       lic_row["licence"],
                        "poids_total_lic": poids_lic,
                        "navire":        navire if 'navire' in dir() else "",
                        "voyage":        voyage if 'voyage' in dir() else "",
                        "pol":           pol if 'pol' in dir() else "",
                        "depart":        depart_d.strftime("%d/%m/%Y") if 'depart_d' in dir() else "",
                        "eta":           eta_prev,
                        "nb_cnt":        nb_cnt if 'nb_cnt' in dir() else 0,
                        "solde_avant":   float(lic_row["solde_reel"]),
                    })
                    st.download_button(
                        "⬇️ Telecharger la confirmation",
                        xlsx_bytes,
                        file_name="CONF_" + str(client_nom).replace(" ", "_") + ".xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
                except Exception as e:
                    st.warning("Template CONF-COMMANDE-TEMPLATE.xlsx introuvable. (" + str(e) + ")")
    st.markdown('</div>', unsafe_allow_html=True)
